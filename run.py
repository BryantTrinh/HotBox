import discord
import json
from discord.ext import commands, tasks
from discord.ui import View, Button, button
from discord import Interaction
import random
from datetime import datetime, timedelta, timezone
import os
from dotenv import load_dotenv
import asyncio
import sys
import re
import traceback
import openpyxl
from io import BytesIO
from openpyxl.styles import numbers, Alignment

# ----------------------------
# Footer Helper
# ----------------------------

def add_embed_footer(embed: discord.Embed) -> discord.Embed:
    footer_line = "*Luna ❀⋆ coded by <@296181275344109568>*"


    if embed.description:
        embed.description = embed.description.replace(f"\n\n{footer_line}", "")


    if embed.description:
        embed.description += f"\n\n{footer_line}"
    else:
        embed.description = footer_line

    return embed


# -------------------------------
# LOAD ENVIRONMENT VARIABLES
# -------------------------------
load_dotenv()
TOKEN = os.getenv("DISCORD_TOKEN")

# -------------------------------
# CONFIG
# -------------------------------

TARGET_CHANNEL_IDS = [1420560553008697474, 1420601193222111233, 1422420786635079701]
COMMAND_PREFIX = "!"
COOLDOWN_SECONDS = 14 * 24 * 60 * 60
DATA_FILE = "loot_data.json"
LOOT_EMOJIS = {
    "Tickets": "🎟️",
    "Bits": "💠",
    "Gold": "💰",
    "Excellent Dust": "✨",
    "Mint Dust": "🍃",
    "Unopened Dye": "🎨"
}

COOLDOWN_BYPASS_USERS = {296181275344109568, 1370076515429253264, 547733449818243084}

LOOT_TABLE = [
    ("Tickets", (1, 5), 5),
    ("Bits", (500, 1000), 30),
    ("Gold", (500, 1000), 15),
    ("Excellent Dust", (5, 15), 23),
    ("Mint Dust", (5, 15), 22),
    ("Unopened Dye", (1, 3), 5)
]

# -------------------------------
# BOT SETUP
# -------------------------------
intents = discord.Intents.default()
intents.message_content = True
intents.guilds = True
intents.members = True

bot = commands.Bot(
    command_prefix=COMMAND_PREFIX,
    intents=intents,
    help_command=None,
    case_insensitive=True,
    application_id=1420292731942731776
)

# -------------------------------
# DATA STORAGE
# -------------------------------
user_cooldowns = {}
user_loot_history = {}

def load_data():
    global user_cooldowns, user_loot_history
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r") as f:
                raw_data = f.read()
                data = json.loads(raw_data)
                
            user_cooldowns = {
                int(k): datetime.fromisoformat(v).astimezone(timezone.utc)
                for k, v in data.get("cooldowns", {}).items()
            }
            user_loot_history = {
                int(k): [
                    (i, v, datetime.fromisoformat(t).astimezone(timezone.utc))
                    for i, v, t in v_list
                ]
                for k, v_list in data.get("history", {}).items()
            }

        except json.JSONDecodeError as e:
            print(f"[WARNING] loot_data.json is corrupted: {e}. Attempting recovery...")
            try:
                recovered_data = None
                for line in raw_data.splitlines():
                    try:
                        recovered_data = json.loads(line)
                        break
                    except:
                        continue

                if recovered_data:
                    user_cooldowns = {
                        int(k): datetime.fromisoformat(v).astimezone(timezone.utc)
                        for k, v in recovered_data.get("cooldowns", {}).items()
                    }
                    user_loot_history = {
                        int(k): [
                            (i, v, datetime.fromisoformat(t).astimezone(timezone.utc))
                            for i, v, t in v_list
                        ]
                        for k, v_list in recovered_data.get("history", {}).items()
                    }
                    print("[INFO] Partial data restored.")
                else:
                    print("[ERROR] Could not recover any data.")
            except Exception as ex:
                print(f"[ERROR] Failed recovery: {ex}")

def save_data():
    data = {
        "cooldowns": {str(k): v.isoformat() for k, v in user_cooldowns.items()},
        "history": {
            str(k): [(i, v, t.isoformat()) for i, v, t in v_list]
            for k, v_list in user_loot_history.items()
        },
    }
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=4)

load_data()

# -------------------------------
# LOOT FUNCTION
# -------------------------------
def roll_loot():
    items = [item for item, _, _ in LOOT_TABLE]
    ranges = {item: rng for item, rng, _ in LOOT_TABLE}
    weights = [chance for _, _, chance in LOOT_TABLE]

    reward_item = random.choices(items, weights=weights, k=1)[0]
    low, high = ranges[reward_item]

    exp = 3
    u = random.random() ** exp
    reward_value = low + int(u * (high - low))

    return reward_item, reward_value

# -------------------------------
# EVENTS
# -------------------------------
@bot.event
async def on_ready():
    # Register persistent views once (for tickets)
    if not getattr(bot, "persistent_views_registered", False):
        from jobboard import MilestoneTicketView
        print("[PersistentViews] Registering MilestoneTicketView globally...")
        bot.add_view(MilestoneTicketView())  # No args; metadata comes from embed footer
        bot.persistent_views_registered = True
        print("[PersistentViews] Registered.")

    # Sync application commands once
    if not getattr(bot, "synced", False):
        bot.tree.copy_global_to(guild=None)
        await bot.tree.sync()
        bot.synced = True
        print("Slash commands synced!")

    print(f"Bot is ready as {bot.user}")


# -------------------------------
# COMMAND HELPERS
# -------------------------------
def find_member_by_name_or_id(guild, query: str):
    query_lower = query.lower()

    if query.startswith("<@") and query.endswith(">"):
        try:
            user_id = int(query.strip("<@!>"))
            return guild.get_member(user_id)
        except ValueError:
            return None

    if query.isdigit():
        return guild.get_member(int(query))

    for member in guild.members:
        if query_lower in member.display_name.lower() or query_lower in member.name.lower():
            return member

    return None

# -------------------------------
# MINI GAME
# -------------------------------
class DoubleOrNothingView(View):
    def __init__(self, ctx, user_id, item, value, timestamp):
        super().__init__(timeout=30)
        self.ctx = ctx
        self.user_id = user_id
        self.item = item
        self.value = value
        self.timestamp = timestamp
        self.interaction_message = None
        self.remaining = 30
        self.ended = False  

    async def start_timer(self):
        try:
            while self.remaining > 0:
                await asyncio.sleep(1)
                self.remaining -= 1
                if self.ended:
                    return
                if self.interaction_message and any(not child.disabled for child in self.children):
                    embed = self.build_embed()
                    embed = add_embed_footer(embed)
                    await self.interaction_message.edit(embed=embed, view=self)

            if not self.ended and self.interaction_message:
                self.ended = True
                for child in self.children:
                    child.disabled = True
                await self.interaction_message.edit(view=self)

                result_text = f"{self.ctx.author.mention} safely kept **{self.value} {self.item}** (Auto Timeout)."
                gif_url = "https://cdn.discordapp.com/attachments/1420560553008697474/1422085281632489514/CFB31F85-BD99-423B-9BE8-7973659FC0C7.gif"

                timeout_embed = discord.Embed(
                    title="Double or Nothing Result",
                    description=result_text + "\n\nCreate a ticket to claim your prize in <#1412934283613700136>",
                    color=0xFFC5D3
                )
                timeout_embed.set_image(url=gif_url)
                timeout_embed = add_embed_footer(timeout_embed)
                await self.interaction_message.edit(embed=timeout_embed, view=None)

                emoji = "⏳"
                label = "Timeout Keep"
                hist_item = f"{emoji} {label} — {self.value} {self.item}"

                user_loot_history.setdefault(self.user_id, []).append(
                    (hist_item, self.value, self.timestamp)
                )
                save_data()

        except Exception as e:
            print(f"[Timer error] {e}")

    def build_embed(self):
        embed = discord.Embed(
            title="🎰 Double or Nothing?",
            description=(
                f"{self.ctx.author.mention}, you won **{self.value} {self.item}!**\n\n"
                f"Do you want to risk it?\n\n"
                f"Pick one: 🎲 (Double) | 🍀 (Keep)\n"
                f"⏳ **Time remaining:** {self.remaining} seconds\n\n"
                f"**Prize:** {self.value} {self.item}"
            ),
            color=0xFFC5D3
        )
        return add_embed_footer(embed)

    async def show_result_embed(self, interaction: Interaction, result_text: str, gif_url: str, outcome_tag: str):
        """
        outcome_tag must be one of:
        🎉 Doubled
        💀 Lost in Double
        🍀 Kept
        ⏳ Timeout Keep
        """
        self.ended = True
        for child in self.children:
            child.disabled = True
        await interaction.message.edit(view=self)

        result_embed = discord.Embed(
            title="Double or Nothing Result",
            description=result_text + (
                "\n\nCreate a ticket to claim your prize in <#1412934283613700136>"
                if self.value > 0 else ""
            ),
            color=0xFFC5D3
        )
        result_embed.set_image(url=gif_url)
        result_embed = add_embed_footer(result_embed)
        await interaction.response.edit_message(embed=result_embed, view=None)

        emoji = outcome_tag.split()[0]
        label = outcome_tag.split(maxsplit=1)[1]

        hist_item = f"{emoji} {label} — {self.value} {self.item}"

        user_loot_history.setdefault(self.user_id, []).append(
            (hist_item, self.value, self.timestamp)
        )
        save_data()

    @button(label="🎲 Double", style=discord.ButtonStyle.success)
    async def double_button(self, interaction: Interaction, button: Button):
        if interaction.user.id != self.user_id:
            return await interaction.response.send_message("❌ This isn’t your lootbox!", ephemeral=True)

        if random.random() < 0.5:
            self.value *= 2
            result_text = f"🎉 {interaction.user.mention} doubled their reward! Now **{self.value} {self.item}**"
            gif_url = "https://cdn.discordapp.com/attachments/1420560553008697474/1422038487569268747/120AEC33-8409-4DDF-9EA3-D6C95DC0D030.gif"
            outcome = "🎉 Doubled"
        else:
            self.value = 0
            result_text = f"💀 {interaction.user.mention} lost it all!"
            gif_url = "https://cdn.discordapp.com/attachments/1420560553008697474/1422040286791733409/IMG_9535.gif"
            outcome = "💀 Lost in Double"

        await self.show_result_embed(interaction, result_text, gif_url, outcome)

    @button(label="🍀 Keep", style=discord.ButtonStyle.secondary)
    async def safe_button(self, interaction: Interaction, button: Button):
        if interaction.user.id != self.user_id:
            return await interaction.response.send_message("❌ This isn’t your lootbox!", ephemeral=True)

        result_text = f"{interaction.user.mention} safely kept **{self.value} {self.item}**."
        gif_url = "https://cdn.discordapp.com/attachments/1420560553008697474/1422085281632489514/CFB31F85-BD99-423B-9BE8-7973659FC0C7.gif"
        outcome = "🍀 Kept"

        await self.show_result_embed(interaction, result_text, gif_url, outcome)


# -------------------------------
# COMMANDS
# -------------------------------
@bot.command(name="open")
async def open_lootbox(ctx):
    try:
        if ctx.channel.id not in TARGET_CHANNEL_IDS:
            return

        user_id = ctx.author.id
        now = datetime.now(timezone.utc)

        member = ctx.guild.get_member(user_id)
        if member is None:
            member = await ctx.guild.fetch_member(user_id)

        if user_id not in COOLDOWN_BYPASS_USERS:
            if not member.premium_since:
                await ctx.send("❌ You must be boosting the server to open a lootbox.")
                return

        if user_id in pending_retries:
            pending_retries.remove(user_id)
        elif user_id not in COOLDOWN_BYPASS_USERS and user_id in user_cooldowns:
            last_open = user_cooldowns[user_id]
            elapsed = (now - last_open).total_seconds()
            if elapsed < COOLDOWN_SECONDS:
                await ctx.send("⏳ You are on cooldown!")
                await check_cooldown(ctx)
                await loot_history(ctx)
                return

        item, value = roll_loot()
        user_cooldowns[user_id] = now
        save_data()

        praying_embed = discord.Embed(
            title="🌸 Praying ♡",
            description=f"{random.choice(list(LOOT_EMOJIS.values()))}   "
                        f"{random.choice(list(LOOT_EMOJIS.values()))}   "
                        f"{random.choice(list(LOOT_EMOJIS.values()))}",
            color=0xFFC5D3
        )
        message = await ctx.send(embed=praying_embed)
        await asyncio.sleep(0.5)

        spin_embed = discord.Embed(title="🎰 Spinning...", description="| 🎲 💎 💰 |", color=0xFFC5D3)
        await message.edit(embed=spin_embed)

        final_emoji = LOOT_EMOJIS[item]

        for _ in range(3):
            reels = [random.choice(list(LOOT_EMOJIS.values())) for _ in range(3)]
            spin_embed.description = f"| {reels[0]}   {reels[1]}   {reels[2]} |"
            await message.edit(embed=spin_embed)
            await asyncio.sleep(0.2)

        spin_embed.description = f"| {final_emoji}   {final_emoji}   {final_emoji} |"
        await message.edit(embed=spin_embed)
        await asyncio.sleep(0.5)

        view = DoubleOrNothingView(ctx, user_id, item, value, now)
        don_embed = view.build_embed()
        don_embed = add_embed_footer(don_embed)
        view.interaction_message = await ctx.send(embed=don_embed, view=view)
        asyncio.create_task(view.start_timer())

    except Exception:
        traceback.print_exc()


@bot.command(name="history")
async def loot_history(ctx, *, query: str = None):
    if query:
        target = find_member_by_name_or_id(ctx.guild, query)
        if not target:
            await ctx.send(f"❌ Could not find a user matching '{query}'.")
            return
    else:
        target = ctx.author

    user_id = target.id
    if user_id not in user_loot_history or len(user_loot_history[user_id]) == 0:
        await ctx.send(f"{target.mention} hasn't opened any lootboxes yet!")
        return

    history = list(reversed(user_loot_history[user_id]))
    items_per_page = 5
    total_pages = (len(history) + items_per_page - 1) // items_per_page
    thumbnail_url = "https://cdn.discordapp.com/attachments/1321372597572599869/1420595192150622409/IMG_9439.gif"

    class LootHistoryView(View):
        def __init__(self):
            super().__init__(timeout=120)
            self.current_page = 0

        async def update_message(self, interaction: Interaction):
            start_index = self.current_page * items_per_page
            end_index = start_index + items_per_page
            page_items = history[start_index:end_index]

            lines = [f"{ts.strftime('%Y-%m-%d %H:%M:%S')}: {item}"
                     for item, value, ts in page_items]

            embed = discord.Embed(
                title=f"{target.display_name}'s Loot History (Page {self.current_page + 1}/{total_pages})",
                description="\n".join(lines),
                color=0xFFDBE5,
            )
            embed.set_thumbnail(url=thumbnail_url)
            embed = add_embed_footer(embed)
            self.previous.disabled = self.current_page == 0
            self.next.disabled = self.current_page == total_pages - 1

            await interaction.response.edit_message(embed=embed, view=self)

        @button(label="◀️", style=discord.ButtonStyle.secondary)
        async def previous(self, interaction: Interaction, button: Button):
            if self.current_page > 0:
                self.current_page -= 1
                await self.update_message(interaction)

        @button(label="▶️", style=discord.ButtonStyle.secondary)
        async def next(self, interaction: Interaction, button: Button):
            if self.current_page < total_pages - 1:
                self.current_page += 1
                await self.update_message(interaction)

    view = LootHistoryView()
    page_items = history[:items_per_page]
    lines = [f"{ts.strftime('%Y-%m-%d %H:%M:%S')}: {item}" for item, value, ts in page_items]

    embed = discord.Embed(
        title=f"{target.display_name}'s Loot History (Page 1/{total_pages})",
        description="\n".join(lines),
        color=0xFFDBE5,
    )
    embed.set_thumbnail(url=thumbnail_url)
    embed = add_embed_footer(embed)
    await ctx.send(embed=embed, view=view)


@bot.command(name="cooldown", aliases=["cd"])
async def check_cooldown(ctx, *, query: str = None):
    now = datetime.now(timezone.utc)

    if query:
        target = find_member_by_name_or_id(ctx.guild, query)
        if not target:
            await ctx.send(f"❌ Could not find a user matching '{query}'.")
            return
    else:
        target = ctx.author

    user_id = target.id
    if user_id not in user_cooldowns:
        await ctx.send(f"{target.mention} can open a lootbox right now!")
        return

    last_open = user_cooldowns[user_id]
    elapsed = (now - last_open).total_seconds()

    if elapsed >= COOLDOWN_SECONDS:
        await ctx.send(f"{target.mention} can open a lootbox right now!")
    else:
        next_time = last_open + timedelta(seconds=COOLDOWN_SECONDS)
        unix_time = int(next_time.timestamp())
        formatted_time = f"<t:{unix_time}:f>"

        embed = discord.Embed(
            title="Cooldown Active ⏳",
            color=0xFFFFFF,
        )
        embed.add_field(
            name="\u200b",
            value=f"{target.mention} can open another chest on: {formatted_time}",
            inline=False,
        )
        embed.add_field(
            name="\u200b",
            value="*Luna ❀⋆ coded by <@296181275344109568>*",
            inline=False,
        )
        await ctx.send(embed=embed)


@bot.command(name="help")
async def help_command(ctx):
    embed = discord.Embed(
        title="<a:hnote3:1420614028514033685> Bot Commands",
        description="Here are all the commands you can use:",
        color=0xF1DBB6
    )
    embed.add_field(name="!open", value="Open a lootbox! Can only be used once every 14 days.", inline=False)
    embed.add_field(name="!history", value="View your lootbox history with paging buttons.", inline=False)
    embed.add_field(name="!cooldown", value="Check how long until you can open your next lootbox.", inline=False)
    embed.add_field(name="!help", value="Show this help message with all available commands.", inline=False)
    embed.add_field(name="!prize", value="This shows list of available prizes.", inline=False)
    embed.set_thumbnail(url="https://cdn.discordapp.com/attachments/1420560553008697474/1420613932288180265/IMG_9442.jpg")
    embed = add_embed_footer(embed)
    await ctx.send(embed=embed)


@bot.command(name="prize")
async def prize_command(ctx):
    embed = discord.Embed(
        title="<a:BunnyBook:1420995026741104804> List of Prizes",
        description="\n",
        color=0xF1DBB6
    )
    embed.add_field(name="Tickets 🎟️", value="\u200b", inline=False)
    embed.add_field(name="Bits 💠", value="\u200b", inline=False)
    embed.add_field(name="Gold 💰", value="\u200b", inline=False)
    embed.add_field(name="Excellent Dust ✨", value="\u200b", inline=False)
    embed.add_field(name="Mint Dust 🍃", value="\u200b", inline=False)
    embed.add_field(name="Unopened Dye 🎨", value="\u200b", inline=False)
    embed.set_thumbnail(
        url="https://cdn.discordapp.com/attachments/1420560553008697474/1420993021972840468/IMG_9468.gif"
    )
    embed = add_embed_footer(embed)
    await ctx.send(embed=embed)


# ------------------------------- 
# Work mechanic 
# ------------------------------- 
MENTION_REGEX = r"<@!?(\d+)>"
USERNAME_REGEX = r"@(\w+)"

def strip_markdown(text):
    if not text:
        return "" 
    text = re.sub(r"[*_~>|]", "", text) 
    text = text.replace("", "")
    text = text.replace("\n", "").strip()
    return text

class LeaderboardView(View):
    def __init__(self, leaderboard_data: list[tuple]):
        """
        leaderboard_data: list of tuples -> [(username, user_id, times_worked), ...]
        """
        super().__init__(timeout=None)
        self.leaderboard_data = leaderboard_data
        self.ALLOWED_USERS = {296181275344109568, 1370076515429253264}

    @button(label="📋 Download Excel", style=discord.ButtonStyle.primary)
    async def download_excel(self, interaction: discord.Interaction, button: Button):
        if interaction.user.id not in self.ALLOWED_USERS:
            return await interaction.response.send_message(
                "❌ You are not allowed to download this file.", ephemeral=True
            )
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Leaderboard"

            ws.append(["Discord User", "User ID", "Times Worked", "Add Command"])

            for username, user_id, times in self.leaderboard_data:
                clean_username = strip_markdown(username)
                command = f"/add name: clanlottery tickets: {times} user: <@{user_id}>"
                ws.append([username, str(user_id), times, command])

                cmd_cell = ws.cell(row=ws.max_row, column=4)
                cmd_cell.number_format = "@"

            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.alignment = Alignment(horizontal="left")

            for col in [3, 4]:
                for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left")

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            file = discord.File(fp=buffer, filename="work_leaderboard.xlsx")
            await interaction.response.send_message(
                "✅ Excel file with RaffleBot commands generated!", file=file, ephemeral=True
            )

        except Exception as e:
            print("Excel generation failed:", e)
            await interaction.response.send_message(
                "❌ Failed to generate Excel file.", ephemeral=True
            )


# -------------------------------------
# WORKED COMMAND
# -------------------------------------
@bot.command(name="worked")
async def worked_command(ctx, days: int = 30):
    ALLOWED_USERS = {1370076515429253264, 296181275344109568}
    WORK_CHANNEL_ID = 1444398245320196117
    TARGET_PHRASE = "your workers have finished their tasks"

    if ctx.author.id not in ALLOWED_USERS:
        return await ctx.send("❌ You do not have permission to use this command.")

    if not (1 <= days <= 90):
        return await ctx.send("❌ Please provide a number between 1 and 90.")

    channel = ctx.guild.get_channel(WORK_CHANNEL_ID)
    if not channel:
        return await ctx.send("❌ Could not find the target work channel.")

    since = datetime.now(timezone.utc) - timedelta(days=days)

    user_counts = {}
    total_scanned = 0
    total_matched = 0

    PROGRESS_BAR_LENGTH = 20
    HEARTBEAT_INTERVAL = 50

    progress_embed = discord.Embed(
        title="🏗️ Scanning Messages...",
        description=f"⏳ Scanning messages from the past {days} days...\nScanned: 0\nMatches: 0",
        color=discord.Color.orange(),
    )
    progress_msg = await ctx.send(embed=progress_embed)

    scanned_since_last_update = 0

    async for message in channel.history(limit=None, oldest_first=False):

        if message.created_at < since:
            break

        total_scanned += 1
        scanned_since_last_update += 1

        if not message.embeds:
            continue

        if message.author.bot:
            for e in message.embeds:
                desc = strip_markdown(e.description.lower() if e.description else "")
                if TARGET_PHRASE.lower() in desc:
                    total_matched += 1
                    original_desc = e.description or ""

                    mentions = re.findall(MENTION_REGEX, original_desc)
                    if mentions:
                        for uid in mentions:
                            uid = int(uid)
                            user_counts[uid] = user_counts.get(uid, 0) + 1
                    else:
                        usernames = re.findall(USERNAME_REGEX, original_desc)
                        for username in usernames:
                            member = discord.utils.find(
                                lambda m: m.name.lower() == username.lower(),
                                ctx.guild.members
                            )
                            if member:
                                user_counts[member.id] = user_counts.get(member.id, 0) + 1

        if scanned_since_last_update >= HEARTBEAT_INTERVAL:
            scanned_since_last_update = 0
            filled = min(PROGRESS_BAR_LENGTH, (total_scanned % 1000) // 50)
            bar = "#" * filled + "-" * (PROGRESS_BAR_LENGTH - filled)

            progress_embed.description = (
                f"⏳ Scanning messages...\n"
                f"**Progress:** [{bar}] Scanned: {total_scanned} messages\n"
                f"**Matches:** {total_matched}"
            )
            await progress_msg.edit(embed=progress_embed)

    bar = "#" * PROGRESS_BAR_LENGTH
    progress_embed.description = (
        f"✅ Scan complete!\n"
        f"Scanned {total_scanned} messages.\n"
        f"Found {total_matched} successful completions."
    )
    await progress_msg.edit(embed=progress_embed)

    if total_matched == 0:
        embed = discord.Embed(
            title="📭 No Work Completions Found",
            description=f"Scanned **{total_scanned}** messages.\nNo matches found.",
            color=discord.Color.red(),
        )
        return await ctx.send(embed=embed)

    filtered_counts = {}
    for user_id, count in user_counts.items():
        if ctx.guild.get_member(user_id):
            filtered_counts[user_id] = count

    user_counts = filtered_counts

    sorted_users = sorted(user_counts.items(), key=lambda x: x[1], reverse=True)

    leaderboard_data = []
    preview_lines = []

    for i, (user_id, count) in enumerate(sorted_users, start=1):
        member = ctx.guild.get_member(user_id)
        username = member.display_name if member else f"<@{user_id}>"

        leaderboard_data.append((username, user_id, count))

        if i <= 10:
            preview_lines.append(f"**#{i}** {username} — {count} times")

    preview_text = "\n".join(preview_lines) or "No valid users found."

    final_embed = discord.Embed(
        title=f"🏆 Work Leaderboard (Past {days} Days)",
        description=(
            f"✅ Scan complete!\n"
            f"Scanned **{total_scanned}** messages.\n"
            f"Found **{total_matched}** completions.\n\n"
            f"{preview_text}"
        ),
        color=discord.Color.green(),
    )

    view = LeaderboardView(leaderboard_data)
    final_embed = add_embed_footer(final_embed)
    await ctx.send(embed=final_embed, view=view)


# -------------------------------
# RETRY MECHANIC
# -------------------------------
RETRY_WHITELIST = {1370076515429253264, 296181275344109568}
pending_retries = set()

@bot.command(name="retry")
async def retry_command(ctx, target: discord.Member):
    """Allows whitelisted users to grant a retry to another user."""
    if ctx.author.id not in RETRY_WHITELIST:
        return await ctx.send("❌ You are not allowed to give retries.")

    pending_retries.add(target.id)
    await ctx.send(f"✅ {target.mention} can now use `!open` again immediately (cooldown bypass once).")


@bot.command(name="inspectkaruta")
async def inspect_karuta(ctx):
    """Inspect the last few messages from the Karuta bot to see embed structure."""
    WORK_CHANNEL_ID = 1284648556849795113
    channel = ctx.guild.get_channel(WORK_CHANNEL_ID)
    if channel is None:
        return await ctx.send("❌ Could not find the target channel.")

    print("\n--- DEBUG: Inspecting last few embeds ---")
    count = 0
    async for message in channel.history(limit=10):
        if message.author.bot and message.embeds:
            print(f"\n[Message {count+1}] From: {message.author} | ID: {message.id}")
            for e in message.embeds:
                print(e.to_dict())
            count += 1
    if count == 0:
        print("No bot messages with embeds found.")
    await ctx.send("✅ Printed last few Karuta embeds to console/logs.")


# -------------------------------
# RUN BOT
# -------------------------------
async def main():
    async with bot:
        await bot.load_extension("jobboard")
        await bot.load_extension("sanriot")
        await bot.load_extension("deal")
        await bot.load_extension("lists")
        await bot.load_extension("daddy")
        await bot.start(TOKEN)

        print("Tree Commands Loaded:", [cmd.name for cmd in bot.tree.get_commands()])


if __name__ == "__main__":
    asyncio.run(main())
