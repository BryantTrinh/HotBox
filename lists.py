import discord
from discord.ext import commands
import json
import os
import openpyxl
from openpyxl.styles import Alignment
from io import BytesIO

LISTS_FILE = "lists.json"


# ----------------------------
# Footer Helper
# ----------------------------
def add_embed_footer(embed: discord.Embed) -> discord.Embed:
    footer_line = "\n\n*Luna ❀⋆ coded by <@296181275344109568>*"

    if embed.description:
        embed.description += footer_line
    else:
        embed.description = footer_line

    return embed


# ---------------------------------
# WHITELIST DECORATOR
# ---------------------------------
WHITELIST = {296181275344109568, 1370076515429253264, 418838937625296898}   # Bry + Luv + mel

def is_whitelisted():
    async def predicate(ctx):
        return ctx.author.id in WHITELIST
    return commands.check(predicate)


# ---------------------------
# Load / Save Functions
# ---------------------------
def load_lists():
    if not os.path.exists(LISTS_FILE):
        with open(LISTS_FILE, "w") as f:
            json.dump({}, f)

    with open(LISTS_FILE, "r") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {}


def save_lists(data):
    with open(LISTS_FILE, "w") as f:
        json.dump(data, f, indent=4)


# ---------------------------
# List Manager Cog
# ---------------------------
class ListManager(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.lists = load_lists()

    def find_member(self, guild, query):
        query = str(query).strip()

        if query.startswith("<@") and query.endswith(">"):
            try:
                return guild.get_member(int(query.strip("<@!>")))
            except:
                pass

        if query.isdigit():
            member = guild.get_member(int(query))
            if member:
                return member

        query_lower = query.lower()
        for m in guild.members:
            if query_lower in m.display_name.lower() or query_lower in m.name.lower():
                return m

        return None


    # ---------------------------
    # Create a new list
    # ---------------------------
    @commands.command(name="createlist")
    @is_whitelisted()
    async def create_list(self, ctx, listname: str):
        listname = listname.lower()

        if listname in self.lists:
            return await ctx.send(f"❌ The list **{listname}** already exists.")

        self.lists[listname] = []
        save_lists(self.lists)

        await ctx.send(f"✅ Created list **{listname}**.")


    # ---------------------------
    # Delete a list
    # ---------------------------
    @commands.command(name="deletelist")
    @is_whitelisted()
    async def delete_list(self, ctx, listname: str):
        listname = listname.lower()

        if listname not in self.lists:
            return await ctx.send("❌ That list does not exist.")

        del self.lists[listname]
        save_lists(self.lists)

        await ctx.send(f"🗑️ Deleted list **{listname}**.")


    # ---------------------------
    # Add a user
    # ---------------------------
    @commands.command(name="listadd")
    @is_whitelisted()
    async def add_to_list(self, ctx, listname: str, *, userquery: str):
        listname = listname.lower()

        if listname not in self.lists:
            return await ctx.send("❌ This list does not exist.")

        member = self.find_member(ctx.guild, userquery)
        if not member:
            return await ctx.send("❌ Could not find that user.")

        if member.id in self.lists[listname]:
            return await ctx.send(f"⚠️ {member.mention} is already in **{listname}**.")

        self.lists[listname].append(member.id)

        self.lists[listname] = sorted(
            self.lists[listname],
            key=lambda uid: (
                ctx.guild.get_member(uid).display_name.lower()
                if ctx.guild.get_member(uid) else "zzzzz"
            )
        )

        save_lists(self.lists)
        await ctx.send(f"➕ Added {member.mention} to **{listname}**.")


    # ---------------------------
    # Remove a user
    # ---------------------------
    @commands.command(name="listremove")
    @is_whitelisted()
    async def remove_from_list(self, ctx, listname: str, *, userquery: str):
        listname = listname.lower()

        if listname not in self.lists:
            return await ctx.send("❌ This list does not exist.")

        member = self.find_member(ctx.guild, userquery)
        if not member:
            return await ctx.send("❌ Could not find that user.")

        if member.id not in self.lists[listname]:
            return await ctx.send(f"⚠️ {member.mention} is not in **{listname}**.")

        self.lists[listname].remove(member.id)
        save_lists(self.lists)

        await ctx.send(f"➖ Removed {member.mention} from **{listname}**.")


    # ---------------------------
    # Show a single list
    # ---------------------------
    @commands.command(name="showlist")
    @is_whitelisted()
    async def show_list(self, ctx, listname: str):
        listname = listname.lower()

        if listname not in self.lists:
            return await ctx.send("❌ That list does not exist.")

        user_ids = self.lists[listname]

        if not user_ids:
            return await ctx.send(f"📭 **{listname}** is empty.")

        sorted_members = sorted(
            user_ids,
            key=lambda uid: (
                ctx.guild.get_member(uid).display_name.lower()
                if ctx.guild.get_member(uid) else "zzzzz"
            )
        )

        lines = []
        for uid in sorted_members:
            m = ctx.guild.get_member(uid)
            if m:
                lines.append(f"• {m.mention} (`{uid}`)")
            else:
                lines.append(f"• <@{uid}> (left server)")

        embed = discord.Embed(
            title=f"📋 List: {listname}",
            description="\n".join(lines),
            color=0x89CFF0
        )
        embed = add_embed_footer(embed)

        await ctx.send(embed=embed)


    # ---------------------------
    # List viewer with buttons
    # ---------------------------
    @commands.command(name="listall")
    @is_whitelisted()
    async def list_lists(self, ctx):
        if not self.lists:
            return await ctx.send("📭 No lists exist yet.")

        list_names = sorted(self.lists.keys())
        total_pages = len(list_names)
        parent = self

        class ListScrollView(discord.ui.View):
            def __init__(self):
                super().__init__(timeout=120)
                self.index = 0

            def build_embed(self):
                listname = list_names[self.index]
                user_ids = parent.lists[listname]

                sorted_members = sorted(
                    user_ids,
                    key=lambda uid: (
                        ctx.guild.get_member(uid).display_name.lower()
                        if ctx.guild.get_member(uid) else "zzzzz"
                    )
                )

                if not sorted_members:
                    users_text = "*(empty)*"
                else:
                    lines = []
                    for uid in sorted_members:
                        m = ctx.guild.get_member(uid)
                        if m:
                            lines.append(f"- {m.mention}")
                        else:
                            lines.append(f"- <@{uid}> (left server)")
                    users_text = "\n".join(lines)

                embed = discord.Embed(
                    title=f"📚 List Viewer ({self.index + 1}/{total_pages})",
                    description=f"**List:** `{listname}`\n\n{users_text}",
                    color=0xFFD1DC,
                )
                embed = add_embed_footer(embed)
                return embed

            async def update(self, interaction):
                embed = self.build_embed()
                await interaction.response.edit_message(embed=embed, view=self)

            @discord.ui.button(label="◀️", style=discord.ButtonStyle.secondary)
            async def previous(self, interaction, button):
                if self.index > 0:
                    self.index -= 1
                    await self.update(interaction)

            @discord.ui.button(label="▶️", style=discord.ButtonStyle.secondary)
            async def next(self, interaction, button):
                if self.index < total_pages - 1:
                    self.index += 1
                    await self.update(interaction)

        view = ListScrollView()
        embed = view.build_embed()
        await ctx.send(embed=embed, view=view)


    # ---------------------------
    # Export to Excel
    # ---------------------------
    @commands.command(name="listexport")
    @is_whitelisted()
    async def export_lists(self, ctx):
        if not self.lists:
            return await ctx.send("📭 No lists exist to export.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lists"

        sorted_listnames = sorted(self.lists.keys())

        col_index = 1

        for listname in sorted_listnames:
            col_letter = openpyxl.utils.get_column_letter(col_index)
            ws[f"{col_letter}1"] = listname

            user_ids = self.lists[listname]
            sorted_users = sorted(
                user_ids,
                key=lambda uid: (
                    ctx.guild.get_member(uid).display_name.lower()
                    if ctx.guild.get_member(uid)
                    else "zzzzz"
                )
            )

            row = 2
            for uid in sorted_users:
                member = ctx.guild.get_member(uid)
                if member:
                    text = f"{member.display_name} ({uid})"
                else:
                    text = f"(left server) ({uid})"

                ws[f"{col_letter}{row}"] = text
                row += 1

            max_length = max(len(str(ws[f"{col_letter}{r}"].value)) for r in range(1, row))
            ws.column_dimensions[col_letter].width = max_length + 2

            col_index += 1

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file = discord.File(buffer, filename="all_lists.xlsx")
        await ctx.send("✅ Exported all lists to Excel!", file=file)


    # ---------------------------
    # Help Command
    # ---------------------------
    @commands.command(name="listhelp")
    @is_whitelisted()
    async def list_help(self, ctx):
        embed = discord.Embed(
            title="📚 List System Commands",
            description="Manage lists, prize tiers, and user groups:",
            color=0xF1DBB6
        )

        embed.add_field(
            name="📁 List Management",
            value=(
                "**!createlist `<name>`** — Create a new list\n"
                "**!deletelist `<name>`** — Delete a list"
            ),
            inline=False,
        )

        embed.add_field(
            name="➕ User Management",
            value=(
                "**!listadd `<list>` `<user>`** — Add user to list\n"
                "**!listremove `<list>` `<user>`** — Remove user from list"
            ),
            inline=False,
        )

        embed.add_field(
            name="📄 Viewing Lists",
            value=(
                "**!showlist `<name>`** — Show users in a list\n"
                "**!listall** — Scroll through using ◀️ ▶️\n"
                "**!listexport** — Export lists to Excel"
            ),
            inline=False,
        )

        embed.add_field(
            name="ℹ️ Help",
            value="**!listhelp** — Show this help menu",
            inline=False
        )

        embed.set_thumbnail(
            url="https://cdn.discordapp.com/emojis/1044329271182995506.gif?size=96&quality=lossless"
        )

        embed.add_field(
        name="\u200b",
        value="*Luna ❀⋆ coded by <@296181275344109568>*",
        inline=False
    )
        await ctx.send(embed=embed)


    @commands.command(name="listrandom")
    @is_whitelisted()
    async def random_pairings(self, ctx, listname: str):
        listname = listname.lower()

        if listname not in self.lists:
            return await ctx.send("❌ That list does not exist.")

        user_ids = self.lists[listname]

        if len(user_ids) == 0:
            return await ctx.send(f"📭 **{listname}** is empty.")

        if len(user_ids) < 2:
            return await ctx.send("⚠️ Not enough users to pair.")

        # If there are an odd number of users, handle it accordingly.
        if len(user_ids) % 2 != 0:
            return await ctx.send(
                f"⚠️ **{listname}** has an odd number of members.\n"
                "Add/remove 1 user before pairing."
            )

        members = []
        for uid in user_ids:
            m = ctx.guild.get_member(uid)
            if m:
                members.append((m.display_name, uid))
            else:
                members.append((f"(left server)", uid))

        import random
        random.shuffle(members)

        pairs = []
        for i in range(len(members)):
            a = members[i]
            b = members[(i + 1) % len(members)]
            pairs.append((a, b))

        lines = []
        for a, b in pairs:
            lines.append(f"**• {a[0]}** (`{a[1]}`)  **➡️**  **{b[0]}** (`{b[1]}`)")

        embed = discord.Embed(
            title=f"🔀 Random Pairing — {listname}",
            description="\n".join(lines),
            color=0xB6F5CA
        )
        embed = add_embed_footer(embed)
        await ctx.send(embed=embed)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{listname} Pairs"

        ws.append(["Partner A", "Partner B"])

        for a, b in pairs:
            ws.append([f"{a[0]} ({a[1]})", f"{b[0]} ({b[1]})"])


        for col in ws.columns:
            max_len = 0
            for cell in col:
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = max_len + 3


        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file = discord.File(buffer, filename=f"{listname}_pairings.xlsx")
        await ctx.send("📄 Download your pairing sheet:", file=file)









# ---------------------------
# Setup
# ---------------------------
async def setup(bot):
    await bot.add_cog(ListManager(bot))
